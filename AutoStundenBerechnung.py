import os
import re
from pathlib import Path

import openpyxl
from openpyxl.styles import Font
from openpyxl.styles.alignment import Alignment

RESET = "\033[0m"
BOLD = "\033[01m"
UNDERLINE = "\033[04m"
BLACK = "\033[30m"
RED = "\033[31m"
GREEN = "\033[32m"
ORANGE = "\033[33m"


class AutoStundenBerechnung:
    def __init__(self, wochenarbeitszeit_soll: float) -> None:
        self.root_path = Path.cwd()
        self.wochenarbeitszeit_soll = wochenarbeitszeit_soll
        os.system("color")

    def get_excellist_path(self) -> list:
        """_summary_

        Returns:
            list: _description_
        """
        xlsm_paths = []
        for xlsm_path in self.root_path.glob("**/*.xl*"):
            pattern = re.search(rf".*KW [\d]*-[\d]*.xl[a-z]+", str(xlsm_path))
            if pattern:
                xlsm_paths.append(xlsm_path)
        if not xlsm_paths:
            raise FileExistsError("Es wurde keine Excellisten gefunden!")
        return sorted(xlsm_paths)

    def lese_ist_wochenarbeitszeit(self, excel_path: Path, excel_sheet: str) -> list:
        """_summary_

        Args:
            excel_path (Path): _description_
            excel_sheet (str, optional): _description_.

        Returns:
            list: _description_
        """
        stundenaufstellung = []
        kalenderwoche = None
        wochenarbeitszeit = None
        try:
            wb = openpyxl.load_workbook(excel_path, data_only=True)
        except Exception:
            raise RuntimeError(f"Bitte Excel {excel_path.name} schließen!")
        max_col = wb[excel_sheet].max_column + 1
        max_row = wb[excel_sheet].max_row + 1
        for i_row in range(1, max_row):
            for i_col in range(1, max_col):
                cell_obj = wb[excel_sheet].cell(row=i_row, column=i_col)
                if cell_obj.value == "Wochenarbeitszeit":
                    wochenarbeitszeit = (
                        wb[excel_sheet].cell(row=i_row + 1, column=i_col).value
                    )
                    kalenderwoche = (
                        wb[excel_sheet].cell(row=i_row, column=i_col - 1).value
                    )
                if kalenderwoche and wochenarbeitszeit:
                    if (
                        dic := {
                            "kalenderwoche": kalenderwoche,
                            "wochenarbeitszeit": wochenarbeitszeit,
                        }
                    ) not in stundenaufstellung:
                        stundenaufstellung.append(dic)
        return stundenaufstellung

    def berechne_wochenarbeitszeit(self, excel_sheet_name: str) -> None:
        """_summary_"""
        index = 2
        print(f"{UNDERLINE}{BOLD}Stunden werden berechnet ...{RESET}")
        # Excel-Header
        wb = openpyxl.Workbook()
        sheet = wb.active
        c1 = sheet.cell(row=1, column=1)
        c1.value = "KW"
        c2 = sheet.cell(row=1, column=2)
        c2.value = "Arbeitsstunden"
        c3 = sheet.cell(row=1, column=3)
        c3.value = "Ueberstunden"
        c4 = sheet.cell(row=1, column=4)
        c4.value = "Gesamte Ueberstunden"
        ueberstunden_liste = []
        for excel_path in self.get_excellist_path():
            zeiten = self.lese_ist_wochenarbeitszeit(excel_path, excel_sheet_name)
            for zeit in zeiten:
                ueberstunden = round(
                    float(zeit["wochenarbeitszeit"])
                    - float(self.wochenarbeitszeit_soll),
                    2,
                )
                ueberstunden_liste.append(ueberstunden)
                if ueberstunden > 0:
                    print(
                        f"{GREEN}Du hast in KW {zeit['kalenderwoche']} "
                        f"-> +{BOLD}{ueberstunden:.2f}h{RESET}{GREEN} "
                        f"zu VIEL gearbeitet.{RESET}"
                    )
                else:
                    print(
                        f"{RED}Du hast in KW {zeit['kalenderwoche']} "
                        f"-> {BOLD}{ueberstunden:.2f}h{RESET}{RED} "
                        f"zu WENIG arbeitet.{RESET}"
                    )
                # Excel-Export
                kw_excel = sheet.cell(row=index, column=1)
                kw_excel.value = zeit["kalenderwoche"]
                stunden_excel = sheet.cell(row=index, column=2)
                stunden_excel.value = zeit["wochenarbeitszeit"]
                uberstundenstunden_excel = sheet.cell(row=index, column=3)
                uberstundenstunden_excel.value = ueberstunden
                index += 1
        # Excel-Export
        total_ueberstunden = sum(ueberstunden_liste)
        gesamte_uberstundenstunden_excel = sheet.cell(row=2, column=4)
        gesamte_uberstundenstunden_excel.value = total_ueberstunden
        # User Ausgabe
        print("-" * 50)
        if total_ueberstunden > 0:
            print(
                f"{GREEN}Deine gesamten Überstunde sind: "
                f"{UNDERLINE}{BOLD}{total_ueberstunden:.2f}h.{RESET}"
            )
        else:
            print(
                f"{RED}Du bist insgesamt mit "
                f"{UNDERLINE}{BOLD}{total_ueberstunden:.2f}"
                f"h{RESET}{RED} im minus.{RESET}"
            )
        print("-" * 50)
        # Excel Settings
        sheet.column_dimensions["A"].width = 10
        sheet.column_dimensions["B"].width = 20
        sheet.column_dimensions["C"].width = 20
        sheet.column_dimensions["D"].width = 30
        sheet.merge_cells(f"D2:D{index-1}")
        sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
        sheet["B1"].alignment = Alignment(horizontal="center", vertical="center")
        sheet["C1"].alignment = Alignment(horizontal="center", vertical="center")
        sheet["D1"].alignment = Alignment(horizontal="center", vertical="center")
        sheet["D2"].alignment = Alignment(horizontal="center", vertical="center")
        sheet["A1"].font = Font(size=14, bold=True)
        sheet["B1"].font = Font(size=14, bold=True)
        sheet["C1"].font = Font(size=14, bold=True)
        sheet["D1"].font = Font(size=14, bold=True)
        sheet["D2"].font = Font(size=14, bold=True, color="00993366")
        # Excel Export in Datei
        try:
            print("Erstellung einer Excelliste ...")
            wb.save(self.root_path / "deineStunden" / "Ueberstunden.xlsx")
            print("Erstellung Fertig")
        except Exception:
            raise RuntimeError("Bitte schließe die Excel-Liste:" "'Ueberstunden.xlsx'")


if __name__ == "__main__":
    asb = AutoStundenBerechnung(wochenarbeitszeit_soll=31.0)
    asb.berechne_wochenarbeitszeit(excel_sheet_name="Stundenaufstellung")
    print("Skript Fertig.")

"""
Install:
python -m pip install --upgrade pip
Run:
py -m PyInstaller --onefile -c AutoStundenBerechnung.py

"""

